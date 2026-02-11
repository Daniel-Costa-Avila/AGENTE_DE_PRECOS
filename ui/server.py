from __future__ import annotations

import os
import shutil
import subprocess
import time
import sys
from dataclasses import dataclass, field
from pathlib import Path
from queue import Queue
from threading import Thread
from uuid import uuid4

from fastapi import FastAPI, Request, UploadFile, File
from fastapi.responses import HTMLResponse, FileResponse, RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook

# ---------------- BASE PATH ----------------

BASE_DIR = Path(__file__).resolve().parent.parent
RUNS_DIR = BASE_DIR / "runs"
DEFAULT_INPUT = BASE_DIR / "input.xlsx"

# ---------------- APP ----------------

app = FastAPI()

# templates
templates = Jinja2Templates(
    directory=str(BASE_DIR / "ui" / "templates")
)

# static files (CSS)
app.mount(
    "/static",
    StaticFiles(directory=str(BASE_DIR / "ui" / "static")),
    name="static"
)


@dataclass
class Job:
    job_id: str
    status: str = "QUEUED"
    created_at: float = field(default_factory=time.time)
    started_at: float | None = None
    finished_at: float | None = None
    input_path: Path | None = None
    output_path: Path | None = None
    error: str | None = None


jobs: dict[str, Job] = {}
job_queue: Queue[str] = Queue()


def _run_job(job: Job) -> None:
    env = os.environ.copy()
    env["PYTHONPATH"] = str(BASE_DIR)
    cmd = [
        sys.executable,
        "-m",
        "App.main",
        "--input",
        str(job.input_path),
        "--output",
        str(job.output_path),
    ]
    subprocess.run(cmd, cwd=str(BASE_DIR), check=True, env=env)


def _worker() -> None:
    while True:
        job_id = job_queue.get()
        job = jobs.get(job_id)
        if job is None:
            job_queue.task_done()
            continue
        job.status = "RUNNING"
        job.started_at = time.time()
        try:
            _run_job(job)
            job.status = "DONE"
        except Exception as exc:
            job.status = "FAILED"
            job.error = f"{type(exc).__name__}: {exc}"
        finally:
            job.finished_at = time.time()
            job_queue.task_done()


@app.on_event("startup")
def startup() -> None:
    RUNS_DIR.mkdir(exist_ok=True)
    worker = Thread(target=_worker, daemon=True)
    worker.start()


# ---------------- HELPERS ----------------

def read_output(path: Path):
    if not path.exists():
        return [], []

    wb = load_workbook(
        path,
        read_only=True,
        data_only=True
    )
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(row)

    wb.close()
    return headers, rows


def _counts() -> dict[str, int]:
    out = {"QUEUED": 0, "RUNNING": 0, "DONE": 0, "FAILED": 0}
    for job in jobs.values():
        if job.status in out:
            out[job.status] += 1
    return out


def _new_job_id() -> str:
    return uuid4().hex[:10]


def _job_payload(job: Job) -> dict:
    return {
        "job_id": job.job_id,
        "status": job.status,
        "created_at": job.created_at,
        "started_at": job.started_at,
        "finished_at": job.finished_at,
        "error": job.error,
    }


# ---------------- ROUTES ----------------

@app.get("/api/health")
def api_health():
    return JSONResponse({"status": "ok"})


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    counts = _counts()
    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "counts": counts,
            "has_default_input": DEFAULT_INPUT.exists(),
        },
    )


@app.post("/run")
def run_agent(file: UploadFile | None = File(default=None)):
    counts = _counts()
    if counts["QUEUED"] + counts["RUNNING"] >= 5:
        return HTMLResponse(
            "Limite de execucoes simultaneas atingido. Tente novamente.",
            status_code=429,
        )
    job_id = _new_job_id()
    job_dir = RUNS_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    input_path = job_dir / "input.xlsx"
    output_path = job_dir / "output.xlsx"

    if file is not None:
        with input_path.open("wb") as f:
            shutil.copyfileobj(file.file, f)
        file.file.close()
    else:
        if not DEFAULT_INPUT.exists():
            return HTMLResponse(
                "Nenhuma planilha enviada e input.xlsx nao encontrado.",
                status_code=400,
            )
        shutil.copy(DEFAULT_INPUT, input_path)

    job = Job(
        job_id=job_id,
        input_path=input_path,
        output_path=output_path,
    )
    jobs[job_id] = job
    job_queue.put(job_id)

    return RedirectResponse(f"/status/{job_id}", status_code=303)


@app.post("/api/run")
def api_run(file: UploadFile | None = File(default=None)):
    counts = _counts()
    if counts["QUEUED"] + counts["RUNNING"] >= 5:
        return JSONResponse(
            {"error": "Limite de execucoes simultaneas atingido."},
            status_code=429,
        )
    job_id = _new_job_id()
    job_dir = RUNS_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    input_path = job_dir / "input.xlsx"
    output_path = job_dir / "output.xlsx"

    if file is not None:
        with input_path.open("wb") as f:
            shutil.copyfileobj(file.file, f)
        file.file.close()
    else:
        if not DEFAULT_INPUT.exists():
            return JSONResponse(
                {"error": "input.xlsx nao encontrado."},
                status_code=400,
            )
        shutil.copy(DEFAULT_INPUT, input_path)

    job = Job(
        job_id=job_id,
        input_path=input_path,
        output_path=output_path,
    )
    jobs[job_id] = job
    job_queue.put(job_id)

    return JSONResponse(
        {
            "job_id": job_id,
            "status_url": f"/status/{job_id}",
            "download_url": f"/download/{job_id}",
        }
    )


@app.get("/status/{job_id}", response_class=HTMLResponse)
def status(request: Request, job_id: str):
    job = jobs.get(job_id)
    if job is None:
        return HTMLResponse("Job nao encontrado.", status_code=404)
    return templates.TemplateResponse(
        "status.html",
        {
            "request": request,
            "job": job,
        },
    )


@app.get("/api/status/{job_id}")
def api_status(job_id: str):
    job = jobs.get(job_id)
    if job is None:
        return JSONResponse({"error": "Job nao encontrado."}, status_code=404)
    return JSONResponse(_job_payload(job))


@app.get("/results/{job_id}", response_class=HTMLResponse)
def results(request: Request, job_id: str):
    job = jobs.get(job_id)
    if job is None:
        return HTMLResponse("Job nao encontrado.", status_code=404)
    if job.status != "DONE" or not job.output_path or not job.output_path.exists():
        return templates.TemplateResponse(
            "status.html",
            {
                "request": request,
                "job": job,
            },
        )
    headers, rows = read_output(job.output_path)
    return templates.TemplateResponse(
        "results.html",
        {
            "request": request,
            "headers": headers,
            "rows": rows,
            "job_id": job_id,
        },
    )


@app.get("/download/{job_id}")
def download(job_id: str):
    job = jobs.get(job_id)
    if job is None or not job.output_path or not job.output_path.exists():
        return HTMLResponse("Arquivo nao encontrado.", status_code=404)
    return FileResponse(
        path=job.output_path,
        filename=f"output_{job_id}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
