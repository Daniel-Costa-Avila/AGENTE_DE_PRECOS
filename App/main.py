from __future__ import annotations

import os
import re
import sys
from pathlib import Path
from typing import Callable, Optional

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from App.collectors.webcontinental import coletar as coletar_webcontinental
from App.collectors.magalu import coletar as coletar_magalu
from App.collectors.zema import coletar as coletar_zema
from App.collectors.madeiramadeira import coletar as coletar_madeiramadeira
from App.collectors.probel import coletar as coletar_probel
from App.collectors.mercadolivre import coletar as coletar_mercadolivre
from App.utils.browser import get_driver


# ---------------- CONFIG ----------------

INPUT_FILE = "input.xlsx"
OUTPUT_FILE = "output.xlsx"

CollectorFn = Callable[..., dict]


# ---------------- RESOLVER COLETOR ----------------

def resolve_collector(link: str) -> Optional[CollectorFn]:
    if not link:
        return None

    l = link.lower().strip()

    if "magazineluiza.com.br" in l or "magalu" in l:
        return coletar_magalu

    if "mercadolivre.com.br" in l:
        return coletar_mercadolivre

    if "webcontinental" in l:
        return coletar_webcontinental

    if "zema.com" in l:
        return coletar_zema

    if "madeiramadeira" in l:
        return coletar_madeiramadeira

    if "probel" in l:
        return coletar_probel

    return None


# ---------------- HELPERS EXCEL ----------------

def find_col(ws: Worksheet, names: set[str]) -> Optional[int]:
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v and str(v).strip().lower() in names:
            return col
    return None


def extract_url_from_cell(cell) -> str:
    if getattr(cell, "hyperlink", None) and cell.hyperlink.target:
        return str(cell.hyperlink.target).strip()

    if not cell.value:
        return ""

    s = str(cell.value).strip()

    m = re.search(r'HYPERLINK\(\s*"([^"]+)"', s, re.IGNORECASE)
    if m:
        return m.group(1).strip()

    if "http://" in s.lower() or "https://" in s.lower():
        return s

    return ""


def _parse_ids(raw: str | None) -> set[str] | None:
    if not raw:
        return None
    ids = {v.strip() for v in raw.split(",") if v.strip()}
    return ids or None


def _get_run_filters() -> tuple[int | None, set[str] | None]:
    max_rows = None
    only_ids = None

    # ENV overrides
    env_limit = os.getenv("LIMIT_ROWS")
    env_ids = os.getenv("ONLY_IDS")
    if env_limit:
        try:
            max_rows = int(env_limit)
        except ValueError:
            max_rows = None
    only_ids = _parse_ids(env_ids)

    # CLI flags: --limit N | --ids A,B
    args = sys.argv[1:]
    for i, arg in enumerate(args):
        if arg in {"--limit", "-l"} and i + 1 < len(args):
            try:
                max_rows = int(args[i + 1])
            except ValueError:
                pass
        if arg in {"--ids", "-i"} and i + 1 < len(args):
            only_ids = _parse_ids(args[i + 1]) or only_ids

    return max_rows, only_ids


def _get_io_paths() -> tuple[Path, Path]:
    input_file = os.getenv("INPUT_FILE") or INPUT_FILE
    output_file = os.getenv("OUTPUT_FILE") or OUTPUT_FILE

    args = sys.argv[1:]
    for i, arg in enumerate(args):
        if arg in {"--input", "-in"} and i + 1 < len(args):
            input_file = args[i + 1]
        if arg in {"--output", "-out"} and i + 1 < len(args):
            output_file = args[i + 1]

    return Path(input_file), Path(output_file)


# ---------------- MAIN ----------------

def main():
    input_path, output_path = _get_io_paths()
    if not input_path.exists():
        raise FileNotFoundError(f"Arquivo {input_path} não encontrado.")

    wb_in = load_workbook(input_path)
    ws_in = wb_in.active

    col_id = find_col(ws_in, {
        "id_produto", "produto", "id", "sku", "codigo",
        "codigo_produto", "código", "product_id"
    }) or 1

    col_titulo = find_col(ws_in, {"titulo", "título", "nome", "descricao", "descrição"})
    col_link = find_col(ws_in, {"link", "url", "href"})

    if col_link is None:
        for c in range(1, ws_in.max_column + 1):
            if extract_url_from_cell(ws_in.cell(row=2, column=c)):
                col_link = c
                break

    if col_link is None:
        raise RuntimeError("Não foi possível localizar a coluna de LINK.")

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Output"

    ws_out.append([
        "id_produto", "titulo", "avista", "pix", "prazo", "status", "link"
    ])

    # -------- Selenium (para canais que usam driver) --------
    driver = get_driver(headless=False)

    # -------- Processar produtos --------
    max_rows, only_ids = _get_run_filters()
    processed = 0

    for row in range(2, ws_in.max_row + 1):
        id_produto = ws_in.cell(row=row, column=col_id).value
        id_key = str(id_produto).strip() if id_produto is not None else ""
        if only_ids and id_key not in only_ids:
            continue
        if max_rows is not None and processed >= max_rows:
            break

        titulo = ws_in.cell(row=row, column=col_titulo).value if col_titulo else None
        link = extract_url_from_cell(ws_in.cell(row=row, column=col_link))

        if not link:
            result = {
                "avista": None,
                "pix": None,
                "prazo": None,
                "status": "LINK AUSENTE"
            }
        else:
            coletor = resolve_collector(link)

            if not coletor:
                result = {
                    "avista": None,
                    "pix": None,
                    "prazo": None,
                    "status": "CANAL NÃO SUPORTADO"
                }
            else:
                try:
                    # 🔑 AJUSTE CRÍTICO AQUI
                    if coletor is coletar_magalu:
                        result = coletor(link)
                    else:
                        result = coletor(driver, link)

                    if not isinstance(result, dict):
                        raise ValueError("Coletor não retornou dict")

                except Exception as e:
                    result = {
                        "avista": None,
                        "pix": None,
                        "prazo": None,
                        "status": f"ERRO NO COLETOR: {type(e).__name__} | {e}"
                    }

        ws_out.append([
            id_produto,
            titulo,
            result.get("avista"),
            result.get("pix"),
            result.get("prazo"),
            result.get("status"),
            link,
        ])
        processed += 1

    wb_out.save(output_path)
    driver.quit()
    total = ws_in.max_row - 1
    if max_rows is not None or only_ids:
        print(
            f"OK — {processed} produtos processados (filtro aplicado). "
            f"Arquivo gerado: {output_path}"
        )
    else:
        print(f"OK — {total} produtos processados. Arquivo gerado: {output_path}")


# ---------------- STATUS ----------------

def write_status(status: str):
    try:
        Path("run_status.txt").write_text(status, encoding="utf-8")
    except Exception:
        pass


if __name__ == "__main__":
    write_status("RUNNING")
    try:
        main()
        write_status("FINISHED")
    except Exception:
        write_status("FAILED")
        raise
